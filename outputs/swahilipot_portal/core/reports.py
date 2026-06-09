from io import BytesIO
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.lib.units import mm
except ImportError:
    canvas = None

from django.http import HttpResponse
from django.contrib.staticfiles import finders


# ── Swahilipot Hub branding constants ─────────────────────────────
BRAND_NAME     = "Swahilipot Hub Foundation"
BRAND_TAGLINE  = "Empowering Youth Through Technology, Arts & Entrepreneurship"
BRAND_ADDRESS  = "Mombasa County Governor's Office, Mombasa"
BRAND_WEBSITE  = "www.swahilipothub.co.ke"
BRAND_EMAIL    = "info@swahilipothub.co.ke"
BRAND_PHONE    = "0114 635505"
BRAND_LOGO_STATIC_PATH = "images/swahilipot-logo.jfif"

if canvas:
    BRAND_COLOR = colors.HexColor("#1e40af")
    BRAND_GOLD  = colors.HexColor("#f59e0b")


def _logo_path():
    return finders.find(BRAND_LOGO_STATIC_PATH)


def _draw_pdf_header(pdf, width, height, title):
    HEADER_H = 92

    pdf.setFillColor(BRAND_COLOR)
    pdf.rect(0, height - HEADER_H, width, HEADER_H, fill=1, stroke=0)

    logo_path = _logo_path()
    text_x = 18 * mm
    LOGO_SIZE = 56

    if logo_path:
        try:
            from reportlab.lib.utils import ImageReader
            img = ImageReader(logo_path)
            logo_y = height - HEADER_H + (HEADER_H - LOGO_SIZE) / 2

            pdf.drawImage(
                img,
                14 * mm, logo_y,
                width=LOGO_SIZE,
                height=LOGO_SIZE,
                preserveAspectRatio=True,
                mask="auto",
            )
            text_x = 14 * mm + LOGO_SIZE + 8
        except Exception:
            pass

    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(text_x, height - 28, BRAND_NAME)

    pdf.setFont("Helvetica", 8)
    pdf.drawString(text_x, height - 42, BRAND_TAGLINE)
    pdf.drawString(text_x, height - 53, f"{BRAND_ADDRESS}")
    pdf.drawString(text_x, height - 63,
                   f"{BRAND_WEBSITE} | {BRAND_EMAIL} | {BRAND_PHONE}")

    pdf.setFillColor(BRAND_GOLD)
    pdf.rect(0, height - HEADER_H - 3, width, 3, fill=1, stroke=0)

    pdf.setFillColor(BRAND_COLOR)
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(18 * mm, height - HEADER_H - 22, title)

    pdf.setFillColor(colors.HexColor("#6b7280"))
    pdf.setFont("Helvetica", 8)
    pdf.drawRightString(
        width - 18 * mm,
        height - HEADER_H - 22,
        f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}",
    )

    pdf.setStrokeColor(BRAND_COLOR)
    pdf.setLineWidth(0.5)
    pdf.line(18 * mm, height - HEADER_H - 28, width - 18 * mm, height - HEADER_H - 28)


def _draw_pdf_footer(pdf, width, page_num):
    pdf.setStrokeColor(colors.HexColor("#e5e7eb"))
    pdf.setLineWidth(0.5)
    pdf.line(18 * mm, 22 * mm, width - 18 * mm, 22 * mm)

    pdf.setFillColor(colors.HexColor("#9ca3af"))
    pdf.setFont("Helvetica", 7)
    pdf.drawString(
        18 * mm,
        14 * mm,
        f"© {datetime.now().year} {BRAND_NAME} · {BRAND_WEBSITE} · {BRAND_EMAIL}",
    )
    pdf.drawRightString(width - 18 * mm, 14 * mm, f"Page {page_num}")


# ── EXCEL EXPORT ────────────────────────────────────────────────
def excel_response(filename, headers, rows):
    if Workbook is None:
        text = "\n".join(
            [",".join(str(h) for h in headers)] +
            [",".join(str(v) for v in row) for row in rows]
        )
        response = HttpResponse(text, content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    blue_fill = PatternFill("solid", fgColor="1e40af")
    gold_fill = PatternFill("solid", fgColor="f59e0b")
    white_font = Font(color="FFFFFF", bold=True, size=12)

    ws.merge_cells("A1:G1")
    ws["A1"] = BRAND_NAME
    ws["A1"].font = white_font
    ws["A1"].fill = blue_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = BRAND_TAGLINE
    ws["A2"].fill = blue_fill

    ws.merge_cells("A3:G3")
    ws["A3"] = f"{BRAND_ADDRESS} | {BRAND_WEBSITE} | {BRAND_EMAIL} | {BRAND_PHONE}"
    ws["A3"].fill = blue_fill

    ws.merge_cells("A4:G4")
    ws["A4"].fill = gold_fill

    ws.merge_cells("A6:G6")
    ws["A6"] = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    ws["A6"].alignment = Alignment(horizontal="right")

    header_fill = PatternFill("solid", fgColor="dbeafe")
    header_font = Font(bold=True, color="1e40af")

    header_row = 8
    thin = Side(style="thin", color="93c5fd")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin, top=thin, left=thin, right=thin)

    alt_fill = PatternFill("solid", fgColor="eff6ff")

    for r_idx, row in enumerate(rows):
        excel_row = header_row + 1 + r_idx
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(
                row=excel_row,
                column=c_idx,
                value="" if val is None else str(val),
            )
            if r_idx % 2 == 0:
                cell.fill = alt_fill
            cell.font = Font(size=9)

    # ✅ FIXED COLUMN WIDTH LOGIC (NO MergedCell ERROR)
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

    stream = BytesIO()
    wb.save(stream)

    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


# ── PDF EXPORT ────────────────────────────────────────────────
def pdf_response(filename, title, headers, rows):
    if canvas is None:
        text = title + "\n" + "\n".join(
            [" | ".join(str(h) for h in headers)] +
            [" | ".join(str(v) for v in row) for row in rows]
        )
        response = HttpResponse(text, content_type="text/plain")
        response["Content-Disposition"] = f'attachment; filename="{filename}.txt"'
        return response

    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.lib.pagesizes import A4, landscape

    stream = BytesIO()
    page_width, page_height = A4

    # Use landscape for wide tables (many columns)
    n_cols = len(headers)
    if n_cols > 5:
        page_width, page_height = landscape(A4)

    doc = SimpleDocTemplate(
        stream,
        pagesize=(page_width, page_height),
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=28 * mm,
        bottomMargin=28 * mm,
        title=title,
    )

    styles = getSampleStyleSheet()

    # Header style for table cells
    header_style = ParagraphStyle(
        "HeaderCell",
        parent=styles["Normal"],
        fontSize=8,
        fontName="Helvetica-Bold",
        textColor=rl_colors.white,
        leading=10,
        wordWrap="CJK",
    )

    # Body style for table cells — wraps long text cleanly
    cell_style = ParagraphStyle(
        "BodyCell",
        parent=styles["Normal"],
        fontSize=7.5,
        fontName="Helvetica",
        textColor=rl_colors.HexColor("#111827"),
        leading=10,
        wordWrap="CJK",
    )

    BRAND_BLUE = BRAND_COLOR
    BRAND_GOLD_COLOR = BRAND_GOLD
    ALT_ROW   = rl_colors.HexColor("#eff6ff")
    WHITE     = rl_colors.white

    # ── Logo + brand block as a nested table ────────────────────────────
    logo_path = _logo_path()
    usable_w = page_width - 36 * mm

    # Build header paragraphs
    brand_ps = ParagraphStyle(
        "Brand", parent=styles["Normal"],
        fontSize=13, fontName="Helvetica-Bold",
        textColor=rl_colors.white,
    )
    sub_ps = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=7.5, fontName="Helvetica",
        textColor=rl_colors.white,
    )
    title_ps = ParagraphStyle(
        "Title", parent=styles["Normal"],
        fontSize=11, fontName="Helvetica-Bold",
        textColor=rl_colors.white, spaceAfter=2,
    )
    gen_ps = ParagraphStyle(
        "Gen", parent=styles["Normal"],
        fontSize=7, fontName="Helvetica",
        textColor=rl_colors.HexColor("#e5e7eb"),
    )

    # Brand info cell
    brand_content = [
        Paragraph(BRAND_NAME, brand_ps),
        Paragraph(BRAND_TAGLINE, sub_ps),
        Paragraph(f"{BRAND_ADDRESS}", sub_ps),
        Paragraph(f"{BRAND_WEBSITE} | {BRAND_EMAIL} | {BRAND_PHONE}", sub_ps),
        Spacer(1, 4),
        Paragraph(title, title_ps),
        Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}", gen_ps),
    ]

    if logo_path:
        from reportlab.platypus import Image as RLImage
        try:
            logo_img = RLImage(logo_path, width=50, height=50)
            header_data = [[logo_img, brand_content]]
            header_col_widths = [56, usable_w - 56]
        except Exception:
            header_data = [[brand_content]]
            header_col_widths = [usable_w]
    else:
        header_data = [[brand_content]]
        header_col_widths = [usable_w]

    header_table = Table(header_data, colWidths=header_col_widths)
    header_table.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), BRAND_BLUE),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING",   (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [BRAND_BLUE]),
    ]))

    # ── Data table ────────────────────────────────────────────────────────
    # Convert headers to Paragraph objects for word-wrap
    table_header_row = [Paragraph(str(h), header_style) for h in headers]

    # Convert data rows to Paragraph objects for clean wrapping
    def fmt_val(v):
        return str(v) if v is not None and v != "" else "—"

    data_rows = []
    for row in rows:
        data_rows.append([Paragraph(fmt_val(v), cell_style) for v in row])

    table_data = [table_header_row] + data_rows

    # Distribute column widths evenly, capped at a max
    col_w = usable_w / max(n_cols, 1)

    data_table = Table(table_data, colWidths=[col_w] * n_cols, repeatRows=1)

    # Build alternating row style
    table_style_cmds = [
        # Header row
        ("BACKGROUND",    (0, 0), (-1, 0), BRAND_BLUE),
        ("TEXTCOLOR",     (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING",    (0, 0), (-1, 0), 8),
        # All cells
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 7.5),
        ("TOPPADDING",    (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("GRID",          (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, ALT_ROW]),
    ]
    data_table.setStyle(TableStyle(table_style_cmds))

    # Gold separator line as a thin table
    gold_bar = Table([[""]], colWidths=[usable_w], rowHeights=[3])
    gold_bar.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BRAND_GOLD_COLOR),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    # ── Footer callback ───────────────────────────────────────────────────
    def add_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setStrokeColor(rl_colors.HexColor("#e5e7eb"))
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(
            18 * mm, 20 * mm,
            page_width - 18 * mm, 20 * mm,
        )
        canvas_obj.setFillColor(rl_colors.HexColor("#9ca3af"))
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.drawString(
            18 * mm, 13 * mm,
            f"© {datetime.now().year} {BRAND_NAME} · {BRAND_WEBSITE} · {BRAND_EMAIL}",
        )
        canvas_obj.drawRightString(
            page_width - 18 * mm, 13 * mm,
            f"Page {doc_obj.page}",
        )
        canvas_obj.restoreState()

    # ── Build PDF ─────────────────────────────────────────────────────────
    elements = [
        header_table,
        gold_bar,
        Spacer(1, 6 * mm),
        data_table,
    ]

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)

    response = HttpResponse(stream.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response